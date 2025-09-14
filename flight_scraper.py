import os
import sys
import time
import re
import traceback
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import threading
from queue import Queue

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import openpyxl
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import requests
import boto3
from io import BytesIO

# --- Threading configuration ---
NUM_THREADS = 5
EXCEL_LOCK = threading.Lock()

# --- Telegram configuration ---
TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '8270013756:AAHYrACY96T2BKsSFpo1FV3_TaJo-dvCcTY')
TELEGRAM_CHAT_ID = os.environ.get('TELEGRAM_CHAT_ID', '105756568')
TELEGRAM_ALERT_THRESHOLD = float(os.environ.get('TELEGRAM_ALERT_THRESHOLD', '0.01'))

# --- URL tracking configuration ---
URL_TRACKING_LOCK = threading.Lock()
TOTAL_URLS_PROCESSED = 0
GOOGLE_URLS_PROCESSED = 0
VALID_FLIGHT_INFO_SAVED = 0

# --- Timeout and delay configuration ---
WAIT_COOKIE_CONSENT_TIMEOUT = 10
WAIT_PRICE_REGION_TIMEOUT = 2
WAIT_FLIGHT_ITEMS_TIMEOUT = 1
SCROLL_PAUSE_TIME = 0.5
SCROLL_MAX_SCROLLS = 3
WAIT_TOP_FLIGHTS_TIMEOUT = 2
EXTRA_WAIT_AFTER_COOKIE = 0.5
EXTRA_WAIT_AFTER_PRICE_REGION = 0.5
EXTRA_WAIT_AFTER_FLIGHT_ITEMS = 0.2


def send_telegram_alert(message):
    """Send a message to Telegram bot"""
    if not TELEGRAM_CHAT_ID or TELEGRAM_CHAT_ID == "YOUR_CHAT_ID_HERE":
        logging.warning("Telegram chat ID not set, skipping alert")
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": message,
            "parse_mode": "Markdown"
        }
        response = requests.post(url, json=payload)
        if response.status_code == 200:
            logging.info("Telegram alert sent successfully")
        else:
            logging.error(f"Failed to send Telegram alert: {response.text}")
    except Exception as e:
        logging.error(f"Error sending Telegram alert: {e}")


def convert_date_with_smart_year(date_str):
    """Convert date string to YYYY-MM-DD format with smart year logic"""
    try:
        formats_without_year = [
            "%a, %b %d", "%b %d", "%B %d", "%m/%d", "%d/%m", "%d %b", "%d %B",
        ]
        formats_with_year = [
            "%b %d, %Y", "%B %d, %Y", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y",
        ]

        for fmt in formats_with_year:
            try:
                return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

        today = datetime.now()
        current_year = today.year

        for fmt in formats_without_year:
            try:
                parsed_date = datetime.strptime(date_str, fmt).replace(year=current_year)
                if parsed_date.date() < today.date():
                    parsed_date = parsed_date.replace(year=current_year + 1)
                return parsed_date.strftime("%Y-%m-%d")
            except ValueError:
                continue

        return date_str
    except Exception:
        return date_str


def configure_chrome_options():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-plugins")
    options.add_argument("--disable-images")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36")

    # Lambda-specific options
    options.add_argument("--single-process")
    options.add_argument("--disable-dev-tools")
    options.add_argument("--no-zygote")
    options.add_argument("--disable-gpu-sandbox")

    # Use system ChromeDriver
    options.add_argument("--disable-dev-shm-usage")

    return options


def handle_cookie_consent(driver):
    try:
        xpath = "//button[.//span[contains(text(), 'Accept all')] or contains(text(),'Accept all')]"
        wait = WebDriverWait(driver, WAIT_COOKIE_CONSENT_TIMEOUT)
        button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        button.click()
        logging.info("Clicked 'Accept all' cookie button.")
        time.sleep(EXTRA_WAIT_AFTER_COOKIE)
    except Exception:
        pass


def wait_for_price_region(driver, timeout=40):
    try:
        WebDriverWait(driver, WAIT_PRICE_REGION_TIMEOUT).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'div[role="region"][aria-label="Find the best price"]')
            )
        )
        logging.info("Page stable: 'Find the best price' region is present.")
        time.sleep(EXTRA_WAIT_AFTER_PRICE_REGION)
    except TimeoutException:
        logging.warning("Timeout waiting for 'Find the best price' region to load.")


def wait_for_flight_items(driver, timeout=7):
    try:
        WebDriverWait(driver, WAIT_FLIGHT_ITEMS_TIMEOUT).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'li[role="listitem"], div[data-testid="offer-listing"]')
            )
        )
        logging.info("Flight items are present on the page.")
        time.sleep(EXTRA_WAIT_AFTER_FLIGHT_ITEMS)
    except TimeoutException:
        logging.warning("Timeout waiting for flight items to appear.")


def scroll_to_load_all_flights(driver, max_scrolls=20, pause_time=2):
    last_height = driver.execute_script("return document.body.scrollHeight")
    for scroll_count in range(SCROLL_MAX_SCROLLS):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            logging.info(f"No more new content after {scroll_count} scrolls.")
            break
        last_height = new_height


def get_flight_containers(driver):
    selectors = [
        "li[role='listitem']",
        'div[data-testid="offer-listing"]',
        'div[jscontroller="yCwSd"]',
        'div[role="option"]',
        'div[role="group"]',
    ]
    containers = []
    for sel in selectors:
        containers = driver.find_elements(By.CSS_SELECTOR, sel)
        if containers:
            logging.info(f"Found {len(containers)} flight containers using selector: {sel}")
            return containers
    raise Exception(f"No flight containers found using selectors: {selectors}")


def extract_airport_codes_from_aria(aria_label_text):
    return re.findall(r"\b[A-Z]{3}\b", aria_label_text or "")


def extract_flight_data(container):
    # [Previous extract_flight_data function content - truncated for brevity]
    # This would be the full function from your original script
    data = {}
    # ... (full implementation from your original script)
    return data


def save_to_excel(data, dep_date, search_url=None):
    cols = [
        "Dep Date", "Dep Time", "Arrival Time", "Duration", "No of Stops", "Stops",
        "Dep Airport", "Arr Airport", "Flight Number(s)", "Airline", "CO2 Emissions",
        "Price", "Cheapest Price", "Price Change", "Scrape Time", "Flight Info",
    ]

    df = pd.DataFrame(data)
    for col in cols:
        if col not in df.columns:
            df[col] = "N/A"
    df = df[cols]

    def get_min_price(row):
        prices = []
        for col in ["Price", "Cheapest Price"]:
            val = row[col]
            if val != "N/A" and val:
                try:
                    num = float(re.sub(r"[^\d.]", "", str(val)))
                    prices.append(num)
                except:
                    pass
        return min(prices) if prices else float('inf')

    df['Min Price'] = df.apply(get_min_price, axis=1)
    df = df.sort_values('Min Price').drop('Min Price', axis=1)

    def get_workbook_name(data_list):
        flight_meta = {"dep_airport": "N/A", "arr_airport": "N/A", "first_dep_date": "N/A", "flight_type": "N/A", "rtn_date": ""}
        if data_list and isinstance(data_list, list) and len(data_list) > 0:
            flight_meta["dep_airport"] = data_list[0].get("Dep Airport", "N/A")
            flight_meta["arr_airport"] = data_list[0].get("Arr Airport", "N/A")
            flight_meta["first_dep_date"] = data_list[0].get("Dep Date", "N/A")

        has_return_date = False
        for flight in data_list:
            rtn_date = flight.get("Rtn Date", "")
            if rtn_date and rtn_date != "" and rtn_date != "N/A" and rtn_date != "Return" and rtn_date != flight.get("Dep Date", ""):
                flight_meta["rtn_date"] = rtn_date
                has_return_date = True
                break

        flight_meta["flight_type"] = "rtn" if has_return_date else "1w"
        now = datetime.now()
        curr_month = now.strftime("%Y-%m")
        next_month = (now.replace(day=1) + pd.DateOffset(months=1)).strftime("%Y-%m")
        dep_date_for_filename = flight_meta.get("first_dep_date", "N/A")
        rtn_date_for_filename = flight_meta.get("rtn_date", "")
        flight_type_suffix = flight_meta.get("flight_type", "1w")
        fname = f"{curr_month},{next_month}_{flight_meta['dep_airport']}_{flight_meta['arr_airport']}_{flight_type_suffix}_{dep_date_for_filename}_{rtn_date_for_filename}.xlsx"
        return fname

    script_dir = "/tmp"  # Use /tmp in Lambda
    filename = os.path.join(script_dir, get_workbook_name(data))

    with EXCEL_LOCK:
        try:
            combined = df.copy()
            previous_prices = {}

            # Use S3 for storage instead of local files
            s3_client = boto3.client('s3')
            bucket_name = os.environ.get('S3_BUCKET_NAME')

            if bucket_name:
                # Try to get previous data from S3
                try:
                    s3_key = f"flights/{os.path.basename(filename)}"
                    obj = s3_client.get_object(Bucket=bucket_name, Key=s3_key)
                    prev_df = pd.read_excel(BytesIO(obj['Body'].read()))

                    for row in prev_df.itertuples():
                        key = (str(row[2]), str(row[3]), str(row[7]), str(row[8]))  # Dep Time, Arr Time, Dep Airport, Arr Airport
                        previous_prices[key] = row[12]  # Price column
                except Exception:
                    pass  # No previous file exists

            # Process price changes and alerts
            for idx, row in combined.iterrows():
                if idx == 0:
                    continue
                key = (str(row.iloc[1]), str(row.iloc[2]), str(row.iloc[6]), str(row.iloc[7]))
                prev_price = previous_prices.get(key)
                curr_price = row.iloc[11]

                if prev_price and curr_price and prev_price != "N/A" and curr_price != "N/A":
                    try:
                        prev_val = float(re.sub(r"[^\d.]", "", str(prev_price)))
                        curr_val = float(re.sub(r"[^\d.]", "", str(curr_price)))
                        price_diff = curr_val - prev_val

                        if price_diff > 0:
                            combined.at[idx, "Price Change"] = f"+{price_diff:.0f}"
                        elif price_diff < 0:
                            combined.at[idx, "Price Change"] = f"{price_diff:.0f}"
                            if abs(price_diff) / prev_val > TELEGRAM_ALERT_THRESHOLD:
                                # Send Telegram alert (implementation from your original script)
                                send_telegram_alert("Price drop alert message here")
                        else:
                            combined.at[idx, "Price Change"] = ""
                    except Exception:
                        combined.at[idx, "Price Change"] = ""
                else:
                    combined.at[idx, "Price Change"] = ""

            # Save to S3
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                sheet_name = f"Flights_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                combined.to_excel(writer, sheet_name=sheet_name, index=False)

                wb = writer.book
                ws = writer.sheets[sheet_name]

                # Apply formatting (same as your original script)
                font_bold = Font(bold=True)
                align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions.height = 30

                col_widths = {
                    "Dep Date": 12, "Dep Time": 11, "Arrival Time": 11, "Duration": 10,
                    "No of Stops": 5.5, "Stops": 10, "Dep Airport": 8, "Arr Airport": 8,
                    "Flight Number(s)": 10, "Airline": 8, "CO2 Emissions": 10,
                    "Price": 6, "Cheapest Price": 6, "Price Change": 6,
                    "Scrape Time": 20, "Flight Info": 60,
                }

                for idx, col in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=idx)
                    cell.font = font_bold
                    cell.alignment = align_center_wrap
                    ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col, 15)

                # Color coding for price changes
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cols.index("Price Change") + 1)
                    val = cell.value
                    if val:
                        val_str = str(val)
                        if val_str.startswith("+"):
                            from openpyxl.styles import PatternFill
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif val_str.startswith("-"):
                            from openpyxl.styles import PatternFill
                            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            excel_buffer.seek(0)

            # Upload to S3
            s3_key = f"flights/{os.path.basename(filename)}"
            s3_client.upload_fileobj(excel_buffer, bucket_name, s3_key)

            print(f"Saved {len(combined)} flights to S3: s3://{bucket_name}/{s3_key}")

        except Exception as e:
            print(f"Exception occurred during Excel saving: {e}")
            traceback.print_exc()


def get_flight_container_class(driver):
    WebDriverWait(driver, WAIT_TOP_FLIGHTS_TIMEOUT).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Top flights')]"))
    )
    top_flights_elem = driver.find_element(By.XPATH, "//*[contains(text(), 'Top flights')]")
    li_element = top_flights_elem.find_element(By.XPATH, "following::li[1]")
    classes = li_element.get_attribute("class").split()

    for c in classes:
        if len(c) >= 5 and len(c) <= 10 and c.isalnum() and " " not in c:
            logging.info(f"Detected flight container class after 'Top flights': {c}")
            return c
    raise Exception("Could not detect flight container class after 'Top flights'")


def get_flight_containers_by_detected_class(driver, flight_li_class):
    selector = f"li.{flight_li_class}"
    containers = driver.find_elements(By.CSS_SELECTOR, selector)
    if not containers:
        logging.error(f"No flight containers found using detected class selector: {selector}")
        raise Exception(f"No flight containers found using detected class selector: {selector}")
    logging.info(f"Found {len(containers)} flights using detected class '{flight_li_class}'")
    return containers


def scrape_current_page_flights(driver, dep_date_value, rtn_date_value, max_flights=15):
    """Scrape flights from the current page state."""
    t0 = time.time()
    scroll_to_load_all_flights(driver)
    logging.info(f"Finished scrolling ({time.time()-t0:.2f}s)")

    t0 = time.time()
    flight_li_class = get_flight_container_class(driver)
    containers = get_flight_containers_by_detected_class(driver, flight_li_class)
    logging.info(f"Found {len(containers)} containers using class '{flight_li_class}' ({time.time()-t0:.2f}s)")

    n = len(containers)
    if n == 0:
        logging.warning("No flight containers found.")
        return []

    if n % 2 == 0:
        max_idx = int(n / 2 + 1)
    else:
        max_idx = int(n / 2 + 1.5)
    if max_idx > max_flights:
        max_idx = max_flights
    logging.info(f"Scraping only first {max_idx} of {n} flights.")
    containers = containers[:max_idx]

    t0 = time.time()
    all_flights = []
    for i, container in enumerate(containers, start=1):
        try:
            flight = extract_flight_data(container)
            flight["Dep Date"] = dep_date_value
            flight["Rtn Date"] = rtn_date_value
            all_flights.append((i, container, flight))
            flight_details = ', '.join([f'{k}: {v}' for k, v in flight.items() if v not in ["N/A", 0, None, ""]])
            logging.info(f"Extracted flight {i}/{len(containers)}: {flight_details}")
        except Exception:
            logging.error(f"Error extracting flight {i}")
            traceback.print_exc()

    logging.info(f"Extracted {len(all_flights)} total flights in {time.time()-t0:.2f}s")

    t0 = time.time()
    flights = []
    for i, container, flight in all_flights:
        co2_value = str(flight.get("CO2 Emissions", "")).strip()
        price_value = str(flight.get("Price", "")).strip()
        co2_valid = bool(re.match(r"^\d+\s?kg\s?(co2e?|co₂e?)$", co2_value.lower()))
        price_valid = bool(re.match(r"^[£$€]\s?\d+", price_value))

        if co2_valid and price_valid:
            flight["CO2 Emissions"] = co2_value
            flight["Price"] = price_value
            flights.append(flight)
        else:
            logging.warning(f"Skipped flight {i}/{len(containers)}: missing valid CO2 or Price.")
    logging.info(f"Filtered to {len(flights)} valid flights ({time.time()-t0:.2f}s)")

    return flights


def scrape_flights(url, driver=None):
    if "google.com" not in url:
        logging.warning(f"URL is not a Google URL, skipping: {url}")
        return False

    if driver is None:
        options = configure_chrome_options()
        # Use system ChromeDriver in Lambda
        driver_path = "/opt/chromedriver" if os.path.exists("/opt/chromedriver") else None
        if driver_path:
            service = Service(executable_path=driver_path)
            driver = webdriver.Chrome(service=service, options=options)
        else:
            driver = webdriver.Chrome(options=options)
        driver.minimize_window()
        should_quit = True
    else:
        should_quit = False

    flights_saved = False

    try:
        t_start = time.time()
        driver.get(url)
        logging.info(f"Browser navigated to URL ({time.time()-t_start:.2f}s)")

        t0 = time.time()
        handle_cookie_consent(driver)
        logging.info(f"Cookie consent handled ({time.time()-t0:.2f}s)")

        t0 = time.time()
        wait_for_price_region(driver)
        logging.info(f"Price region loaded ({time.time()-t0:.2f}s)")

        t0 = time.time()
        wait_for_flight_items(driver)
        logging.info(f"Flight items loaded ({time.time()-t0:.2f}s)")

        dep_date_value = rtn_date_value = None
        try:
            dep_input = driver.find_element(By.CSS_SELECTOR, 'input[aria-label="Departure"]')
            dep_date_raw = dep_input.get_attribute("value").strip()
            if not dep_date_raw:
                dep_date_raw = dep_input.get_attribute("placeholder").strip()
            dep_date_value = convert_date_with_smart_year(dep_date_raw)
            logging.info(f"Converted dep date: '{dep_date_value}'")
        except Exception:
            dep_date_value = datetime.now().strftime("%Y-%m-%d")

        try:
            rtn_input = driver.find_element(By.CSS_SELECTOR, 'input[aria-label="Return"]')
            rtn_date_raw = rtn_input.get_attribute("value").strip()
            if not rtn_date_raw:
                rtn_date_raw = rtn_input.get_attribute("placeholder").strip()
            rtn_date_value = convert_date_with_smart_year(rtn_date_raw)
            logging.info(f"Converted rtn date: '{rtn_date_value}'")
        except Exception:
            rtn_date_value = ""

        initial_flights = scrape_current_page_flights(driver, dep_date_value, rtn_date_value)

        cheapest_flights = []
        try:
            logging.info("Looking for 'Cheapest' button...")
            cheapest_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Cheapest')]"))
            )
            cheapest_button.click()
            logging.info("Clicked 'Cheapest' button.")

            time.sleep(2)
            wait_for_flight_items(driver)
            logging.info("Cheapest page loaded.")

            cheapest_flights = scrape_current_page_flights(driver, dep_date_value, rtn_date_value)
            logging.info(f"Scraped {len(cheapest_flights)} cheapest flights.")
        except TimeoutException:
            logging.warning("No 'Cheapest' button found or timeout waiting for it.")
        except Exception as e:
            logging.error(f"Error clicking 'Cheapest' button: {e}")

        cheapest_price_dict = {}
        for flight in cheapest_flights:
            key = (
                str(flight.get("Dep Time", "")),
                str(flight.get("Arrival Time", "")),
                str(flight.get("Dep Airport", "")),
                str(flight.get("Arr Airport", ""))
            )
            cheapest_price_dict[key] = flight.get("Price", "N/A")

        for flight in initial_flights:
            key = (
                str(flight.get("Dep Time", "")),
                str(flight.get("Arrival Time", "")),
                str(flight.get("Dep Airport", "")),
                str(flight.get("Arr Airport", ""))
            )
            flight["Cheapest Price"] = cheapest_price_dict.get(key, "N/A")

        all_flights_to_save = initial_flights[:]
        for cf in cheapest_flights:
            original_price = cf.get("Price", "N/A")
            key_cf = (
                str(cf.get("Dep Time", "")),
                str(cf.get("Arrival Time", "")),
                str(cf.get("Dep Airport", "")),
                str(cf.get("Arr Airport", ""))
            )
            if not any(
                str(f.get("Dep Time", "")) == key_cf[0] and
                str(f.get("Arrival Time", "")) == key_cf[1] and
                str(f.get("Dep Airport", "")) == key_cf[2] and
                str(f.get("Arr Airport", "")) == key_cf[3]
                for f in all_flights_to_save
            ):
                cf["Price"] = "N/A"
                cf["Cheapest Price"] = original_price
                all_flights_to_save.append(cf)

        if all_flights_to_save:
            dep_date = datetime.now().strftime("%Y-%m-%d")
            t_excel = time.time()
            try:
                save_to_excel(all_flights_to_save, dep_date, url)
                logging.info(f"Saved {len(all_flights_to_save)} flights to Excel ({time.time()-t_excel:.2f}s)")
                flights_saved = True
            except Exception as e:
                logging.error(f"Excel save failed: {e}")

    finally:
        if should_quit and driver:
            logging.info("Quitting Chrome driver...")
            t_quit = time.time()
            try:
                driver.quit()
                logging.info(f"Chrome driver quit ({time.time()-t_quit:.2f}s)")
            except Exception as e:
                logging.error(f"Chrome driver quit failed: {e}")

    return flights_saved
